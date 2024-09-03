import win32com.client as win32
import pandas as pd 
import numpy as np
import os
# import jpype
import shutil
# import asposecells
from tkinter import *
from openpyxl import load_workbook
# from asposecells.api import wb
from tkinter import filedialog as fd
from tkinter.scrolledtext import ScrolledText

root = Tk()

root.title('AMI - SMOC SPOT')
root.geometry('500x600')
# root.config(bg = 'azureblue')

global md, excl, path, exp01, exp09, exp11, exp51, bcdf, newmpath

# function to pyupload template - activated when button is clicked
def gettemp():
    global template
    
    # to open file directory for choosing file
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    Label(root, text = filenames[0]).place(x = 185, y  = 162)
    
    template = filenames[0]
    return template

# function to upload meterdata
def getmd():
    global md, path, totunits
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    Label(root, text = filenames[0]).place(x = 185, y  = 82)
    
    tots = 0
    totunits = []
    sets = []
    names = []
    path = filenames[0]
    setsdf = pd.DataFrame()
    allsets = pd.DataFrame()
    
    # reads meterdata file into a dataframe
    md = pd.read_excel(path, sheet_name = 'Sheet1') 
    md = md[['MR Unit','Portion','Installat.','Equipment']] 
    md = md.rename(columns = {'Equipment' : 'Meter ID'}) 
    md['MR Unit'] = md['MR Unit'].astype(str) 
    md['Installat.'] = md['Installat.'].astype(str)
    
    # get column of stations
    md['Unit'] = md['MR Unit'].str.slice(stop = 3)
    totunits = md['Unit'].drop_duplicates().tolist()
    
    # split by 10ks
    for i in totunits:
        tempdf = md.loc[(md['Unit'] == i), 'Installat.'].reset_index(drop = True)
        tots += len(tempdf)

        if tots > 15000: # doesnt add current MRU
            sets = [x if len(sets) != 0 else i for x in sets]
            if len(setsdf) == 0: setsdf = tempdf

            allsets = pd.concat([allsets, setsdf], axis = 1)
            names.append(','.join(sets))

            # reset total & set, take current on
            tots = len(tempdf)
            sets = []
            setsdf = tempdf

        else:
            setsdf = tempdf if len(setsdf) == 0 else pd.concat([setsdf, tempdf], ignore_index = True)

        sets.append(i)
        
    allsets = pd.concat([allsets, setsdf], axis = 1)
    names.append(','.join(sets))
    
    # to overcome XLSX issue
    shutil.copy(path, 'temp.xlsx')
    shutil.copy('temp.xlsx', path.split('.')[0] + '.xlsx')

    os.remove('temp.xlsx')
    path = path.split('.')[0] + '.xlsx'

    print(allsets)

    allsets.columns = names

    # display first column in 2nd textbox
    textbox.delete('1.0', END)
    textbox.insert(END, '\n'.join(allsets[names[0]].astype(str).tolist()))
    total.config(text = 'Paste in SQVI (' + str(len(allsets[names[0]].tolist())) + ' meters)')
    
    writer = pd.ExcelWriter(path, mode = 'a', engine = 'openpyxl')
    allsets.to_excel(writer, sheet_name = 'Split', index = False)
    writer.close()
    
    return md, path

# function to upload exclusion file
def getexcl():
    global excl
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    Label(root, text = filenames[0]).place(x = 185, y  = 42)
    
    # reading file and only taking columns meter id and remarks
    exclpath = filenames[0]
    excl = pd.read_excel(exclpath)
    excl['Remarks'] = 'Check'
    excl = excl[['Meter ID','Remarks']]
    
    return excl

def getseg():
    global bcdf
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    Label(root, text = filenames[0]).place(x = 185, y  = 202)
    
    bcrm = filenames[0]
    bcdf = pd.read_excel(bcrm)
    return bcdf

def expfile():

    global exp01, exp09, exp11, exp51
    name = fd.askdirectory()
    
    files = [f for f in os.listdir(name)]
    
    for f in files:
        print(f)
        if '01' in str(f): exp01 = pd.read_csv(name + '\\' + f)
        elif '09' in str(f): exp09 = pd.read_csv(name + '\\' + f)
        elif '11' in str(f): exp11 = pd.read_csv(name + '\\' + f)
        elif '51' in str(f): exp51 = pd.read_csv(name + '\\' + f)

    exp01 = exp01.rename(columns = {'METERID' : 'Meter ID', 'FINALREAD' : 'Reg 01 (kWh)','FINALREADSTATUS' : 'Reading Status', 'AVERAGECONSUMPTION_NAME':'Avg. Consumption' })
    exp01 = exp01.loc[exp01['Reading Status'] == 'VAL', ['Meter ID','Reg 01 (kWh)','Reading Status','Avg. Consumption']]
    exp01['Reg 01 (kWh)'] = exp01['Reg 01 (kWh)'].apply(np.floor).astype('Int64')

    # merge exports with meterdata dataframe
    exp09 = formatexp(exp09, 'Reg 09 (kWh)')
    exp11 = formatexp(exp11, 'Reg 11 (kWh)')
    exp51 = formatexp(exp51, 'Reg 51 (kWh)')
    
    Label(root, text = name).place(x = 185, y  = 122)
    
def formatexp(df, reg): #for exp9,11,51
    
    global md
    
    df = df.rename(columns = {'METERID' : 'Meter ID', 'FINALREAD' : reg})
    df = df.loc[df['FINALREADSTATUS'] == 'VAL',['Meter ID',reg]]
    df[reg] = df[reg].apply(np.floor).astype(int)

    return df

def resize(df, name, writer):
    for column in df:
        width = max(df[column].astype(str).map(len).max(), len(column))
        col_idx = df.columns.get_loc(column)
        writer.sheets[name].set_column(col_idx, col_idx, width)
    
def genmd():
    global md
    
    folder = '\\'.join(path.split('/')[:-1])
    spot = path.split('/')[-1].split('.')[0] 

    newmpath = folder + '\\' + spot + '_MD.xlsx'
    writer = pd.ExcelWriter(newmpath, engine = 'xlsxwriter')
    
    # merge exports with meterdata dataframe
    md = pd.merge(left = md, right = exp01, how = 'left')
    md = pd.merge(left = md, right = exp09, how = 'left')
    md = pd.merge(left = md, right = exp11, how = 'left')
    md = pd.merge(left = md, right = exp51, how = 'left')

    # checks for incomplete register
    md = md[['MR Unit','Portion','Installat.','Meter ID','Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)','Reading Status','Avg. Consumption']]    
    md.loc[(pd.isnull(md['Reg 01 (kWh)'])) | (pd.isnull(md['Reg 09 (kWh)'])) | (pd.isnull(md['Reg 11 (kWh)'])) | (pd.isnull(md['Reg 51 (kWh)'])), 'Reading Status'] = '#N/A'
    md.loc[(md['Reading Status'] != 'VAL'), ['Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)']] = np.nan

    md.to_excel(writer, sheet_name = 'Meter Data', index = False)

    # make another dataframe with reg 51 > 2 & not kwh forwarded
    reg51 = md.loc[(md['Reg 51 (kWh)'] >= 2) & (md['Avg. Consumption'] == 'kWh received'), ['Meter ID','Reg 51 (kWh)','Reg 01 (kWh)','Avg. Consumption']]
    check = reg51.loc[(reg51['Meter ID'].str.startswith('HOL')) | (reg51['Meter ID'].str.startswith('EDM')) | (reg51['Meter ID'].str.startswith('SAG100')) | (reg51['Meter ID'].str.startswith('SAG101'))]
    check = pd.merge(left = check, right = excl, how = 'left')
    check.to_excel(writer, sheet_name = 'Check', index = False)
    
    chprint = check.loc[(check['Remarks'] == 'Check'), ['Meter ID', 'Reg 01 (kWh)']]
    
    resize(check, 'Check', writer)
    resize(md, 'Meter Data', writer)
    
    writer.close()
    
    textbox2.delete('1.0', END)
    if len(chprint) > 0:   
        textbox2.insert(END, 'Please check meter(s) on BCRM')
        textbox2.insert(END, '\n') 
        textbox2.insert(END, chprint) 
    else:
        textbox2.insert(END, 'No meters to check!')

def genspot():
    global bcdf,  path, totunits
    
    units = []
    
    folder = '\\'.join(path.split('/')[:-1])
    spot = path.split('/')[-1].split('.')[0] 
    newmpath = folder + '\\' + spot + '_MD.xlsx'
    
    newmd = pd.read_excel(newmpath, sheet_name = 'Meter Data')
    newmd['MR Unit'] = newmd['MR Unit'].astype(str)
    newmd['Installat.'] = newmd['Installat.'].astype(str)
    
    folder = '\\'.join(path.split('/')[:-1])
    spot = path.split('/')[-1].split('.')[0] 
    
    bcdf = bcdf.drop_duplicates()
    bcdf = bcdf.sort_values(['MR Unit', 'Sequence Number'], ascending = [True, True])
    bcdf = bcdf.reset_index()

    bcdf['MR Unit'] = bcdf['MR Unit'].astype(str)
    bcdf['Installat.'] = bcdf['Installat.'].astype(str)
    bcdf['Contract Account'] = bcdf['Contract Account'].astype(str)
    bcdf['Meter Reading Date'] = bcdf['Meter Reading Date'].astype(str)
    
    ind = bcdf['Meter Reading Date'][5].split('-')[1]

    bcdf = bcdf.drop(columns = ['Meter Reading Date','Valid fr.','Valid to','IS'])
            
    bcdf['Unit'] = bcdf['MR Unit'].str.slice(stop = 3)
    units = bcdf['Unit'].drop_duplicates().tolist()
    
    textbox2.delete('1.0', END)
    
    for i in units:
        md2 = newmd[newmd['MR Unit'].str.startswith(i)] 
        bcdf2 = bcdf[bcdf['MR Unit'].str.startswith(i)]  

        # get meter ID
        val = pd.merge(left = bcdf2, right = md2, how = 'left', on = 'Installat.')
        val = val.drop(columns =['MR Unit_y','Portion_y'])
        val = val.rename(columns = {'MR Unit_x' : 'MR Unit','Portion_x' : 'Portion'})
        
        # reorder columns
        val['Remarks'] = ''
        val = val[['MR Unit','Contract Account','Sequence Number','Portion','Installat.','Meter ID','Address','Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)','Remarks','Reading Status']]
        val = val.drop_duplicates()
        
        val['Installat.'] = val['Installat.'].astype(str)
        val['Contract Account'] = val['Contract Account'].astype(str)
        val[['Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)']] = val[['Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)']].astype(str) 
        
        # get rid of decimal
        val.loc[(val['Reg 01 (kWh)'].str.contains('.')), 'Reg 01 (kWh)'] =  val['Reg 01 (kWh)'].str.split(pat = '.').str[0] 
        val.loc[(val['Reg 09 (kWh)'].str.contains('.')), 'Reg 09 (kWh)'] = val['Reg 09 (kWh)'].str.split(pat = '.').str[0] 
        val.loc[(val['Reg 11 (kWh)'].str.contains('.')), 'Reg 11 (kWh)'] = val['Reg 11 (kWh)'].str.split(pat = '.').str[0] 
        val.loc[(val['Reg 51 (kWh)'].str.contains('.')), 'Reg 51 (kWh)'] = val['Reg 51 (kWh)'].str.split(pat = '.').str[0] 
        
        # clean up non-val rows
        val.loc[(val['Reading Status'] != 'VAL'), ['Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)']] = '#N/A'
        val.loc[(val['Reading Status'] != 'VAL'), 'Remarks'] = 'Unsuccessful Reading from System'
        val.loc[(val['Reading Status'] != 'VAL'), 'Reading Status'] = '#N/A'
        
        month = ['JAN','FEB','MAR','APR','MAY','JUNE','JULY','AUG','SEPT','OCT','NOV','DEC']
        
        # will go into same directory as spot folder
        dest = folder + '\\' + spot + ' 6' + i + ' ' + month[int(ind)-1] + '.xlsx' 
        shutil.copy(template, dest)

        writer = pd.ExcelWriter(dest, mode = 'a', engine = 'openpyxl')
        writer.book = load_workbook(dest)
        
        for j in writer.book.sheetnames:
            if j != 'Checklist' and j != 'CHECKLIST':
                writer.book.remove(writer.book[j])
        
        val.to_excel(writer, sheet_name = 'VAL', index = False)
        writer.close()

        # save as binary
        # jpype.startJVM() #accessing java
        # wbook = wb(dest)
        # wbook.save(dest.split('.')[0] + '.xlsb')
        # jpype.shutdownJVM()
        
        xl = win32.Dispatch('Excel.Application')
        wbb = xl.Workbooks.Open(Filename = dest, ReadOnly = True)
        wbb.SaveAs(Filename = dest.split('.')[0] + '.xlsb', FileFormat = 50)
        wbb.Close(False) #close excel without saving
        xl.Quit()
        
        os.remove(dest)

        
        valid = val.loc[val['Reading Status'] == 'VAL']
        noval = val.loc[val['Reading Status'] != 'VAL']
        
        textbox2.insert(END, 'Station: ' + str(i))
        textbox2.insert(END, '\n\n')        
        textbox2.insert(END, 'Valid: ' + str(len(valid)))
        textbox2.insert(END, '\n')
        textbox2.insert(END, 'Not Valid: ' + str(len(noval)))
        textbox2.insert(END, '\n')
        textbox2.insert(END, 'Total: ' + str(len(val)))
        textbox2.insert(END, '\n')
        textbox2.insert(END, '\n')
        
        textbox2.insert(END, str(len(val)) + ' Accounts')
        textbox2.insert(END, '\n')
        textbox2.insert(END, str(len(valid)) + ' Accounts')
        textbox2.insert(END, '\n')
        textbox2.insert(END, str(len(noval)) + ' Accounts')
        textbox2.insert(END, '\n')
        textbox2.insert(END, '-------------------------')
        textbox2.insert(END, '\n')

    # textbox2.delete('1.0', END)
    # textbox2.insert(END, '\n'.join(allsets[names[0]].tolist()))

    # output split column
    tots = 0
    outdf = pd.DataFrame()
    # outunits = [i for i in totunits if i not in units] #wrong

    # must check all files in folder and see if unit is there or not
    # files = []

    # for i in totunits:
    #     if any(i in x for x in os.listdir())
    print(totunits)

    big = [j for j in os.listdir(folder) if os.path.getsize(folder + '//' + j) > 10000]
    outunits = [i for i in totunits if any(i in x for x in big) == False]
    print(outunits)
    # test this !!!!!!!!!!!!!
    
    for i in outunits:
        md = newmd.loc[(newmd['MR Unit'].str.startswith(i)),'Installat.'].reset_index(drop = True)
        tots += len(md)

        if tots < 15000: outdf = md if len(outdf) == 0 else pd.concat([outdf, md], ignore_index = True)
        else: break

    textbox.delete('1.0', END)
    textbox.insert(END, '\n'.join(outdf.tolist()))

    if len(outunits) == 0:     total.config(text = 'Paste in SQVI')
    else: total.config(text = 'Paste in SQVI (' + str(len(outdf)) + ' meters)')
    out.config(text = 'SPOT Files: ')
    
    
def test():
    global totunits, path
    print(totunits)
    folder = '\\'.join(path.split('/')[:-1])

    big = [j for j in os.listdir(folder) if j[-4] == 'xlsb' and os.path.getsize(folder + '//' + j) > 10000]
    outunits = [i for i in totunits if any(i in x for x in big) == False]
    print(outunits)
        
def exitt(event):
    root.quit()


# Uploads
Button(root, text = 'Template SPOT', command = gettemp).place(x = 44, y = 160)
Button(root, text = 'Meter Data BCRM', command = getmd).place(x = 30, y = 80)
Button(root, text = 'Exclusion File', command = getexcl).place(x = 50, y = 40)
Button(root, text = 'Segregated Data', command = getseg).place(x = 37, y = 200)
Button(root, text = 'MyCloud Folder', command = expfile).place(x = 37, y = 120)

Label(root, text = 'Path: ').place(x = 145, y  = 42)
Label(root, text = 'Path: ').place(x = 145, y  = 82)
Label(root, text = 'Path: ').place(x = 145, y  = 122)
Label(root, text = 'Path: ').place(x = 145, y  = 162)
Label(root, text = 'Path: ').place(x = 145, y  = 202)

total = Label(root, text = '')
total.place(x = 35, y = 295)

out = Label(root, text = '')
out.place(x = 255, y = 295)

# b4 = Button(root, text = 'Next', command = nextt)
# b3 = Button(root, text = 'Back', command = back)
# b4.place(x = 67, y = 625)
# b3.place(x = 30, y = 625)

# b3["state"] = DISABLED
# b4["state"] = DISABLED

textvar = StringVar()
textbox = ScrolledText(root, height = 15, width = 25)
textbox.place(x = 35, y = 315)

textvar2 = StringVar()
textbox2 = ScrolledText(root, height = 15, width = 25)
textbox2.place(x = 255, y = 315)

Button(root, text = 'Generate Meter Data', command = genmd).place(x = 125, y = 250) #genmd
Button(root, text = 'Generate SPOT Files', command = genspot).place(x = 255, y = 250)

Label(root, text = 'SAB').place(x = 470, y = 580)

root.resizable(True, False) 
root.bind('<Escape>', exitt)
root.mainloop()