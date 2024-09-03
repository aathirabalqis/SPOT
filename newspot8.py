import pandas as pd 
import numpy as np
import os
import shutil
from tkinter import *
from openpyxl import load_workbook
from tkinter import filedialog as fd
from tkinter.scrolledtext import ScrolledText

root = Tk()

root.title('AMI - SMOC SPOT')
root.geometry('500x500')
# root.config(bg = 'azureblue')

global md, excl, path, exp01, exp09, exp11, exp51, bcdf, newmpath

# function to upload template - activated when button is clicked
def gettemp():
    global template
    
    # to open file directory for choosing file
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    Label(root, text = filenames[0]).place(x = 185, y  = 162)
    
    template = filenames[0]
    return template

# function to upload meterdata
def getmd():
    global md, path
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    Label(root, text = filenames[0]).place(x = 185, y  = 82)
    
    tots = 0
    units = []
    sets = []
    names = []
    path = filenames[0]
    setsdf = pd.DataFrame()
    allsets = pd.DataFrame()
    
    # reads meterdata file into a dataframe
    md = pd.read_excel(path, sheet_name = 'Sheet1') 
    md = md[['MR Unit','Portion','Installat.','Equipment']] 
    md = md.rename(columns = {'Equipment' : 'Meter ID'}) 
    md['MR Unit'] = md['MR Unit'].astype(str) 
    md['Installat.'] = md['Installat.'].astype(str)
    
    # get column of stations
    md['Unit'] = md['MR Unit'].str.slice(stop = 3)
    units = md['Unit'].drop_duplicates().tolist()
    
    # split by 10ks
    for i in units:
        tempdf = md.loc[(md['Unit'] == i), 'Installat.'].reset_index(drop = True)
        tots += len(tempdf)

        if tots > 15000: # doesnt add current MRU
            sets = [x if len(sets) != 0 else i for x in sets]
            if len(setsdf) == 0: setsdf = tempdf

            allsets = pd.concat([allsets, setsdf], axis = 1)
            names.append(','.join(sets))

            # reset total & set, take current on
            tots = len(tempdf)
            sets = []
            setsdf = tempdf

        else:
            setsdf = tempdf if len(setsdf) == 0 else pd.concat([setsdf, tempdf], ignore_index = True)

        sets.append(i)
        
    allsets = pd.concat([allsets, setsdf], axis = 1)
    names.append(','.join(sets))
    
    # to overcome XLSX issue
    shutil.copy(path, 'temp.xlsx')
    shutil.copy('temp.xlsx', path.split('.')[0] + '.xlsx')

    os.remove('temp.xlsx')
    path = path.split('.')[0] + '.xlsx'

    allsets.columns = names
    
    writer = pd.ExcelWriter(path, mode = 'a', engine = 'openpyxl')
    allsets.to_excel(writer, sheet_name = 'Split', index = False)
    writer.close()
    
    return md, path

# function to upload exclusion file
def getexcl():
    global excl
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    Label(root, text = filenames[0]).place(x = 185, y  = 42)
    
    # reading file and only taking columns meter id and remarks
    exclpath = filenames[0]
    excl = pd.read_excel(exclpath)
    excl['Remarks'] = 'Check'
    excl = excl[['Meter ID','Remarks']]
    
    return excl

def getseg():
    global bcdf
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    Label(root, text = filenames[0]).place(x = 185, y  = 202)
    
    bcrm = filenames[0]
    bcdf = pd.read_excel(bcrm)
    return bcdf

def expfile():

    global exp01, exp09, exp11, exp51
    name = fd.askdirectory()
    
    files = [f for f in os.listdir(name)]
    
    for f in files:
        print(f)
        if '01' in str(f): exp01 = pd.read_csv(name + '\\' + f)
        elif '09' in str(f): exp09 = pd.read_csv(name + '\\' + f)
        elif '11' in str(f): exp11 = pd.read_csv(name + '\\' + f)
        elif '51' in str(f): exp51 = pd.read_csv(name + '\\' + f)
    
    # path01 = name + '/export01.csv'
    # path09 = name + '/export09.csv'
    # path11 = name + '/export11.csv'
    # path51 = name + '/export51.csv'
    
    # #read export files
    # exp01 = pd.read_csv(path01)
    # exp09 = pd.read_csv(path09)
    # exp11 = pd.read_csv(path11)
    # exp51 = pd.read_csv(path51)

    exp01 = exp01.rename(columns = {'METERID' : 'Meter ID', 'FINALREAD' : 'Reg 01 (kWh)','FINALREADSTATUS' : 'Reading Status', 'AVERAGECONSUMPTION_NAME':'Avg. Consumption' })
    exp01 = exp01.loc[exp01['Reading Status'] == 'VAL', ['Meter ID','Reg 01 (kWh)','Reading Status','Avg. Consumption']]
    exp01['Reg 01 (kWh)'] = exp01['Reg 01 (kWh)'].apply(np.floor).astype('Int64')

    # merge exports with meterdata dataframe
    exp09 = formatexp(exp09, 'Reg 09 (kWh)')
    exp11 = formatexp(exp11, 'Reg 11 (kWh)')
    exp51 = formatexp(exp51, 'Reg 51 (kWh)')
    
    Label(root, text = name).place(x = 185, y  = 122)
    
def formatexp(df, reg): #for exp9,11,51
    
    global md
    
    df = df.rename(columns = {'METERID' : 'Meter ID', 'FINALREAD' : reg})
    df = df.loc[df['FINALREADSTATUS'] == 'VAL',['Meter ID',reg]]
    df[reg] = df[reg].apply(np.floor).astype(int)

    return df

def resize(df, name, writer):
    for column in df:
        width = max(df[column].astype(str).map(len).max(), len(column))
        col_idx = df.columns.get_loc(column)
        writer.sheets[name].set_column(col_idx, col_idx, width)
    
def genmd():
    global md
    
    folder = '\\'.join(path.split('/')[:-1])
    spot = path.split('/')[-1].split('.')[0] 

    newmpath = folder + '\\' + spot + '_MD.xlsx'
    writer = pd.ExcelWriter(newmpath, engine = 'xlsxwriter')
    
    # merge exports with meterdata dataframe
    md = pd.merge(left = md, right = exp01, how = 'left')
    md = pd.merge(left = md, right = exp09, how = 'left')
    md = pd.merge(left = md, right = exp11, how = 'left')
    md = pd.merge(left = md, right = exp51, how = 'left')

    # checks for incomplete register
    md = md[['MR Unit','Portion','Installat.','Meter ID','Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)','Reading Status','Avg. Consumption']]    
    md.loc[(pd.isnull(md['Reg 01 (kWh)'])) | (pd.isnull(md['Reg 09 (kWh)'])) | (pd.isnull(md['Reg 11 (kWh)'])) | (pd.isnull(md['Reg 51 (kWh)'])), 'Reading Status'] = '#N/A'
    md.loc[(md['Reading Status'] != 'VAL'), ['Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)']] = np.nan
    
    

    md.to_excel(writer, sheet_name = 'Meter Data', index = False)

    # make another dataframe with reg 51 > 2 & not kwh forwarded
    reg51 = md.loc[(md['Reg 51 (kWh)'] >= 2) & (md['Avg. Consumption'] == 'kWh received'), ['Meter ID','Reg 51 (kWh)','Reg 01 (kWh)','Avg. Consumption']]
    check = reg51.loc[(reg51['Meter ID'].str.startswith('HOL')) | (reg51['Meter ID'].str.startswith('EDM')) | (reg51['Meter ID'].str.startswith('SAG100')) | (reg51['Meter ID'].str.startswith('SAG101'))]
    check = pd.merge(left = check, right = excl, how = 'left')
    check.to_excel(writer, sheet_name = 'Check', index = False)
    
    chprint = check.loc[(check['Remarks'] == 'Check'), ['Meter ID', 'Reg 01 (kWh)']]
    
    resize(check, 'Check', writer)
    resize(md, 'Meter Data', writer)
    
    writer.close()
    
    textbox.delete('1.0', END)
    if len(chprint) > 0:   
        textbox.insert(END, 'Please check meter(s) on BCRM')
        textbox.insert(END, '\n') 
        textbox.insert(END, chprint) 
    else:
        textbox.insert(END, 'No meters to check!')

def genspot():
    global bcdf,  path
    
    units = []
    
    folder = '\\'.join(path.split('/')[:-1])
    spot = path.split('/')[-1].split('.')[0] 
    newmpath = folder + '\\' + spot + '_MD.xlsx'
    
    newmd = pd.read_excel(newmpath, sheet_name = 'Meter Data')
    newmd['MR Unit'] = newmd['MR Unit'].astype(str)
    newmd['Installat.'] = newmd['Installat.'].astype(str)
    
    folder = '\\'.join(path.split('/')[:-1])
    spot = path.split('/')[-1].split('.')[0] 
    
    bcdf = bcdf.drop_duplicates()
    bcdf = bcdf.sort_values(['MR Unit', 'Sequence Number'], ascending = [True, True])
    bcdf = bcdf.reset_index()

    bcdf['MR Unit'] = bcdf['MR Unit'].astype(str)
    bcdf['Installat.'] = bcdf['Installat.'].astype(str)
    bcdf['Contract Account'] = bcdf['Contract Account'].astype(str)
    bcdf['Meter Reading Date'] = bcdf['Meter Reading Date'].astype(str)
    
    # print(bcdf['Meter Reading Date'])
    
    # print(bcdf['Meter Reading Date'][5])
    # print(bcdf['Meter Reading Date'][5].split('-'))
    ind = bcdf['Meter Reading Date'][5].split('-')[1]

    
    bcdf = bcdf.drop(columns = ['Meter Reading Date','Valid fr.','Valid to','IS'])
    
    # for i in bcdf['MR Unit']:
        # if i[:3] not in units:
            # units.append(i[:3])
            
    bcdf['Unit'] = bcdf['MR Unit'].str.slice(stop = 3)
    units = bcdf['Unit'].drop_duplicates().tolist()
    
    textbox.delete('1.0', END)
    
    for i in units:
        md2 = newmd[newmd['MR Unit'].str.startswith(i)] 
        bcdf2 = bcdf[bcdf['MR Unit'].str.startswith(i)]  

        # get meter ID
        val = pd.merge(left = bcdf2, right = md2, how = 'left', on = 'Installat.')
        val = val.drop(columns =['MR Unit_y','Portion_y'])
        val = val.rename(columns = {'MR Unit_x' : 'MR Unit','Portion_x' : 'Portion'})
        
        # reorder columns
        val['Remarks'] = ''
        val = val[['MR Unit','Contract Account','Sequence Number','Portion','Installat.','Meter ID','Address','Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)','Remarks','Reading Status']]
        val = val.drop_duplicates()
        
        val['Installat.'] = val['Installat.'].astype(str)
        val['Contract Account'] = val['Contract Account'].astype(str)
        val[['Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)']] = val[['Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)']].astype(str) 
        
        print(val['Reg 01 (kWh)'])
        # get rid of decimal
        val.loc[(val['Reg 01 (kWh)'].str.contains('.')), 'Reg 01 (kWh)'] =  val['Reg 01 (kWh)'].str.split(pat = '.').str[0] 
        val.loc[(val['Reg 09 (kWh)'].str.contains('.')), 'Reg 09 (kWh)'] = val['Reg 09 (kWh)'].str.split(pat = '.').str[0] 
        val.loc[(val['Reg 11 (kWh)'].str.contains('.')), 'Reg 11 (kWh)'] = val['Reg 11 (kWh)'].str.split(pat = '.').str[0] 
        val.loc[(val['Reg 51 (kWh)'].str.contains('.')), 'Reg 51 (kWh)'] = val['Reg 51 (kWh)'].str.split(pat = '.').str[0] 
        
        print(val['Reg 01 (kWh)'])
        
        # clean up non-val rows
        val.loc[(val['Reading Status'] != 'VAL'), ['Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)']] = '#N/A'
        val.loc[(val['Reading Status'] != 'VAL'), 'Remarks'] = 'Unsuccessful Reading from System'
        val.loc[(val['Reading Status'] != 'VAL'), 'Reading Status'] = '#N/A'
        
        month = ['JAN','FEB','MAR','APR','MAY','JUNE','JULY','AUG','SEPT','OCT','NOV','DEC']
        
        # will go into same directory as spot folder
        dest = folder + '\\' + spot + ' 6' + i + ' ' + month[int(ind)-1] + '.xlsx' 
        shutil.copy(template, dest)

        writer = pd.ExcelWriter(dest, mode = 'a', engine = 'openpyxl')
        writer.book = load_workbook(dest)
        
        for j in writer.book.sheetnames:
            if j != 'Checklist' and j != 'CHECKLIST':
                writer.book.remove(writer.book[j])
        # writer.book.remove(writer.book['METER ID'])
        
        val.to_excel(writer, sheet_name = 'VAL', index = False)
        
        # resize(val, 'VAL', writer)
        writer.close()
        
        
        # writer2 = pd.ExcelWriter(dest, engine = 'xlsxwriter')
        # resize(val, 'VAL',writer2)
        # writer2.close()
        
        valid = val.loc[val['Reading Status'] == 'VAL']
        noval = val.loc[val['Reading Status'] != 'VAL']
        
        textbox.insert(END, 'Station: ' + str(i))
        textbox.insert(END, '\n')        
        textbox.insert(END, 'Valid: ' + str(len(valid)))
        textbox.insert(END, '\n')
        textbox.insert(END, 'Not Valid: ' + str(len(noval)))
        textbox.insert(END, '\n')
        textbox.insert(END, 'Total: ' + str(len(val)))
        textbox.insert(END, '\n')
        textbox.insert(END, '\n')
        
        textbox.insert(END, str(len(val)) + ' Accounts')
        textbox.insert(END, '\n')
        textbox.insert(END, str(len(valid)) + ' Accounts')
        textbox.insert(END, '\n')
        textbox.insert(END, str(len(noval)) + ' Accounts')
        textbox.insert(END, '\n')
        textbox.insert(END, '--------------------------------------------')
        textbox.insert(END, '\n')

        
def exitt(event):
    root.quit()


# Uploads
Button(root, text = 'Template SPOT', command = gettemp).place(x = 44, y = 160)
Button(root, text = 'Meter Data BCRM', command = getmd).place(x = 30, y = 80)
Button(root, text = 'Exclusion File', command = getexcl).place(x = 50, y = 40)
Button(root, text = 'Segregated Data', command = getseg).place(x = 37, y = 200)
Button(root, text = 'MyCloud Folder', command = expfile).place(x = 37, y = 120)

Label(root, text = 'Path: ').place(x = 145, y  = 42)
Label(root, text = 'Path: ').place(x = 145, y  = 82)
Label(root, text = 'Path: ').place(x = 145, y  = 122)
Label(root, text = 'Path: ').place(x = 145, y  = 162)
Label(root, text = 'Path: ').place(x = 145, y  = 202)

# b4 = Button(root, text = 'Next', command = nextt)
# b3 = Button(root, text = 'Back', command = back)
# b4.place(x = 67, y = 625)
# b3.place(x = 30, y = 625)

# b3["state"] = DISABLED
# b4["state"] = DISABLED

textvar = StringVar()
textbox = ScrolledText(root, height = 11, width = 45)
textbox.place(x = 65, y = 290)

Button(root, text = 'Generate Meter Data', command = genmd).place(x = 125, y = 250)
Button(root, text = 'Generate SPOT Files', command = genspot).place(x = 255, y = 250)

Label(root, text = 'SAB').place(x = 470, y = 480)

root.resizable(True, False) 
root.bind('<Escape>', exitt)
root.mainloop()