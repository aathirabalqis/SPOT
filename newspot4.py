import pandas as pd 
import numpy as np
import shutil
from tkinter import *
from openpyxl import load_workbook
from tkinter import filedialog as fd
from tkinter.scrolledtext import ScrolledText

# what if total installation exceeds limit? - make new sheet to separate installation into columns
# meterdata file needs to be named as the correct spot
# resize columns 
# center all columns


# when displaying meter ID to check on BCRM, display reg 1 too

root = Tk()

root.title('AMI - SMOC SPOT')
root.geometry('500x500')

global md, excl, path, exp01, exp09, exp11, exp51, bcdf, newmpath

def gettemp():
    global template
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    Label(root, text = filenames[0]).place(x = 185, y  = 162)
    
    template = filenames[0]
    return template

def getmd():
    global md, path
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    Label(root, text = filenames[0]).place(x = 185, y  = 82)
    
    path = filenames[0]
    
    md = pd.read_excel(path, sheet_name = 'Sheet1')
    # md = md.drop(columns = ['Advanced Metering System','Advanced Meter Capability Grp (AMCG)','Reg.'])
    md = md[['MR Unit','Portion','Installat.','Equipment']]
    md = md.rename(columns = {'Equipment' : 'Meter ID'})
    md['MR Unit'] = md['MR Unit'].astype(str)
    md['Installat.'] = md['Installat.'].astype(str)
    
    # print(md)
    
    return md, path

def getexcl():
    global excl
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    Label(root, text = filenames[0]).place(x = 185, y  = 42)
    
    exclpath = filenames[0]
    excl = pd.read_excel(exclpath)
    excl['Remarks'] = 'Check'
    excl = excl[['Meter ID','Remarks']]
    
    return excl

def getseg():
    global bcdf
    filenames = fd.askopenfilenames(filetypes=[("Text files","*.xlsx")])
    Label(root, text = filenames[0]).place(x = 185, y  = 202)
    
    bcrm = filenames[0]
    bcdf = pd.read_excel(bcrm)
    return bcdf

def expfile():

    global exp01, exp09, exp11, exp51
    name = fd.askdirectory()
    
    path01 = name + '/export01.csv'
    path09 = name + '/export09.csv'
    path11 = name + '/export11.csv'
    path51 = name + '/export51.csv'
    
    #read export files
    exp01 = pd.read_csv(path01)
    exp09 = pd.read_csv(path09)
    exp11 = pd.read_csv(path11)
    exp51 = pd.read_csv(path51)

    exp01 = exp01.rename(columns = {'METERID' : 'Meter ID', 'FINALREAD' : 'Reg 01 (kWh)','FINALREADSTATUS' : 'Reading Status', 'AVERAGECONSUMPTION_NAME':'Avg. Consumption' })
    exp01 = exp01.loc[exp01['Reading Status'] == 'VAL', ['Meter ID','Reg 01 (kWh)','Reading Status','Avg. Consumption']]
    exp01['Reg 01 (kWh)'] = exp01['Reg 01 (kWh)'].apply(np.floor).astype('Int64')

    # merge exports with meterdata dataframe
    exp09 = formatexp(exp09, 'Reg 09 (kWh)')
    exp11 = formatexp(exp11, 'Reg 11 (kWh)')
    exp51 = formatexp(exp51, 'Reg 51 (kWh)')
    
    Label(root, text = name).place(x = 185, y  = 122)
    
def formatexp(df, reg): #for exp9,11,51
    
    global md
    
    df = df.rename(columns = {'METERID' : 'Meter ID', 'FINALREAD' : reg})
    df = df.loc[df['FINALREADSTATUS'] == 'VAL',['Meter ID',reg]]
    df[reg] = df[reg].apply(np.floor).astype(int)

    return df

def resize(df, name, writer):
    for column in df:
        width = max(df[column].astype(str).map(len).max(), len(column))
        col_idx = df.columns.get_loc(column)
        writer.sheets[name].set_column(col_idx, col_idx, width)
    
def genmd():
    global md
    
    folder = '\\'.join(path.split('/')[:-1])
    spot = path.split('/')[-1].split('.')[0] 

    newmpath = folder + '\\' + spot + '_MD.xlsx'
    writer = pd.ExcelWriter(newmpath, engine = 'xlsxwriter')
    
    # merge exports with meterdata dataframe
    md = pd.merge(left = md, right = exp01, how = 'left')
    md = pd.merge(left = md, right = exp09, how = 'left')
    md = pd.merge(left = md, right = exp11, how = 'left')
    md = pd.merge(left = md, right = exp51, how = 'left')

    # checks for incomplete register
    md = md[['MR Unit','Portion','Installat.','Meter ID','Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)','Reading Status','Avg. Consumption']]    
    md.loc[(pd.isnull(md['Reg 01 (kWh)'])) | (pd.isnull(md['Reg 09 (kWh)'])) | (pd.isnull(md['Reg 11 (kWh)'])) | (pd.isnull(md['Reg 51 (kWh)'])), 'Reading Status'] = '#N/A'
    md.loc[(md['Reading Status'] != 'VAL'), ['Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)']] = np.nan
    
    

    md.to_excel(writer, sheet_name = 'Meter Data', index = False)

    # make another dataframe with reg 51 > 2 & not kwh forwarded
    reg51 = md.loc[(md['Reg 51 (kWh)'] >= 2) & (md['Avg. Consumption'] == 'kWh received'), ['Meter ID','Reg 51 (kWh)','Reg 01 (kWh)','Avg. Consumption']]
    check = reg51.loc[(reg51['Meter ID'].str.startswith('HOL')) | (reg51['Meter ID'].str.startswith('EDM')) | (reg51['Meter ID'].str.startswith('SAG100')) | (reg51['Meter ID'].str.startswith('SAG101'))]
    check = pd.merge(left = check, right = excl, how = 'left')
    check.to_excel(writer, sheet_name = 'Check', index = False)
    
    chprint = check.loc[(check['Remarks'] == 'Check'), ['Meter ID', 'Reg 01 (kWh)']]
    
    resize(check, 'Check', writer)
    resize(md, 'Meter Data', writer)
    
    writer.close()
    
    textbox.delete('1.0', END)
    if len(chprint) > 0:   
        textbox.insert(END, 'Please check meter(s) on BCRM')
        textbox.insert(END, '\n') 
        textbox.insert(END, chprint) 
    else:
        textbox.insert(END, 'No meters to check!')

def genspot():
    global bcdf,  path
    
    units = []
    
    folder = '\\'.join(path.split('/')[:-1])
    spot = path.split('/')[-1].split('.')[0] 
    newmpath = folder + '\\' + spot + '_MD.xlsx'
    
    newmd = pd.read_excel(newmpath, sheet_name = 'Meter Data')
    newmd['MR Unit'] = newmd['MR Unit'].astype(str)
    newmd['Installat.'] = newmd['Installat.'].astype(str)
    
    folder = '\\'.join(path.split('/')[:-1])
    spot = path.split('/')[-1].split('.')[0] 

    bcdf['MR Unit'] = bcdf['MR Unit'].astype(str)
    bcdf['Installat.'] = bcdf['Installat.'].astype(str)
    bcdf['Contract Account'] = bcdf['Contract Account'].astype(str)
    bcdf['Meter Reading Date'] = bcdf['Meter Reading Date'].astype(str)
    
    print(bcdf['Meter Reading Date'][5])
    print(bcdf['Meter Reading Date'][5].split('-'))
    ind = bcdf['Meter Reading Date'][5].split('-')[1]

    
    bcdf = bcdf.drop(columns = ['Meter Reading Date','Valid fr.','Valid to','IS'])
    
    for i in bcdf['MR Unit']:
        if i[:3] not in units:
            units.append(i[:3])
    
    textbox.delete('1.0', END)
    
    for i in units:
        md2 = newmd[newmd['MR Unit'].str.startswith(i)] 
        bcdf2 = bcdf[bcdf['MR Unit'].str.startswith(i)]  

        # get meter ID
        val = pd.merge(left = bcdf2, right = md2, how = 'left', on = 'Installat.')
        val = val.drop(columns =['MR Unit_y','Portion_y'])
        val = val.rename(columns = {'MR Unit_x' : 'MR Unit','Portion_x' : 'Portion'})
        
        # reorder columns
        val['Remarks'] = ''
        val = val[['MR Unit','Contract Account','Sequence Number','Portion','Installat.','Meter ID','Address','Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)','Remarks','Reading Status']]
        val = val.drop_duplicates()
        
        val['Installat.'] = val['Installat.'].astype(str)
        val['Contract Account'] = val['Contract Account'].astype(str)
        val[['Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)']] = val[['Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)']].astype(str) 
        
        # get rid of decimal
        val['Reg 01 (kWh)'] = val['Reg 01 (kWh)'].str[:-2]
        val['Reg 09 (kWh)'] = val['Reg 09 (kWh)'].str[:-2]
        val['Reg 11 (kWh)'] = val['Reg 11 (kWh)'].str[:-2]
        val['Reg 51 (kWh)'] = val['Reg 51 (kWh)'].str[:-2]

        # clean up non-val rows
        val.loc[(val['Reading Status'] != 'VAL'), ['Reg 01 (kWh)','Reg 09 (kWh)','Reg 11 (kWh)','Reg 51 (kWh)']] = '#N/A'
        val.loc[(val['Reading Status'] != 'VAL'), 'Remarks'] = 'Unsuccessful Reading from System'
        val.loc[(val['Reading Status'] != 'VAL'), 'Reading Status'] = '#N/A'
        
        month = ['JAN','FEB','MAR','APR','MAY','JUNE','JULY','AUG','SEPT','OCT','NOV','DEC']
        
        # will go into same directory as spot folder
        dest = folder + '\\' + spot + ' 6' + i + ' ' + month[int(ind)-1] + '.xlsx' 
        shutil.copy(template, dest)

        writer = pd.ExcelWriter(dest, mode = 'a', engine = 'openpyxl')
        writer.book = load_workbook(dest)
        
        for j in writer.book.sheetnames:
            if j != 'Checklist' and j != 'CHECKLIST':
                writer.book.remove(writer.book[j])
        # writer.book.remove(writer.book['METER ID'])
        
        val.to_excel(writer, sheet_name = 'VAL', index = False)
        
        # resize(val, 'VAL', writer)
        writer.close()
        
        
        # writer2 = pd.ExcelWriter(dest, engine = 'xlsxwriter')
        # resize(val, 'VAL',writer2)
        # writer2.close()
        
        valid = val.loc[val['Reading Status'] == 'VAL']
        noval = val.loc[val['Reading Status'] != 'VAL']
        
        textbox.insert(END, 'Station: ' + str(i))
        textbox.insert(END, '\n')        
        textbox.insert(END, 'Total Accounts: ' + str(len(val)))
        textbox.insert(END, '\n')
        textbox.insert(END, 'Reading from MDMS: ' + str(len(valid)))
        textbox.insert(END, '\n')
        textbox.insert(END, 'No Reading: ' + str(len(noval)))
        textbox.insert(END, '\n')
        textbox.insert(END, '\n')
        
        textbox.insert(END, str(len(val)) + ' Accounts')
        textbox.insert(END, '\n')
        textbox.insert(END, str(len(valid)) + ' Accounts')
        textbox.insert(END, '\n')
        textbox.insert(END, str(len(noval)) + ' Accounts')
        textbox.insert(END, '\n')
        textbox.insert(END, '--------------------------------------------')
        textbox.insert(END, '\n')

        
def exitt(event):
    root.quit()


# Uploads
Button(root, text = 'Template SPOT', command = gettemp).place(x = 44, y = 160)
Button(root, text = 'Meter Data BCRM', command = getmd).place(x = 30, y = 80)
Button(root, text = 'Exclusion File', command = getexcl).place(x = 50, y = 40)
Button(root, text = 'Segregated Data', command = getseg).place(x = 37, y = 200)
Button(root, text = 'MyCloud Folder', command = expfile).place(x = 37, y = 120)

Label(root, text = 'Path: ').place(x = 145, y  = 42)
Label(root, text = 'Path: ').place(x = 145, y  = 82)
Label(root, text = 'Path: ').place(x = 145, y  = 122)
Label(root, text = 'Path: ').place(x = 145, y  = 162)
Label(root, text = 'Path: ').place(x = 145, y  = 202)

textvar = StringVar()
textbox = ScrolledText(root, height = 11, width = 45)
textbox.place(x = 65, y = 290)

Button(root, text = 'Generate Meter Data', command = genmd).place(x = 125, y = 250)
Button(root, text = 'Generate SPOT Files', command = genspot).place(x = 255, y = 250)

Label(root, text = 'SAB').place(x = 470, y = 480)

root.resizable(True, False) 
root.bind('<Escape>', exitt)
root.mainloop()